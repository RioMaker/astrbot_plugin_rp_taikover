"""把 Excel/CSV 内容表导入 resource/content.json。

表头：类型、内容、RP下限、RP上限、启用、备注
RP 范围留空表示不限制，边界值包含在范围内。
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable


PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from rp_core import CONTENT_KEYS, ContentItem, ContentStore, parse_optional_rp  # noqa: E402


TYPE_ALIASES = {
    "今日签": "fortune_texts",
    "签": "fortune_texts",
    "fortune": "fortune_texts",
    "fortune_texts": "fortune_texts",
    "幸运色": "colors",
    "颜色": "colors",
    "color": "colors",
    "colors": "colors",
    "宜": "advice_do",
    "今日宜": "advice_do",
    "do": "advice_do",
    "advice_do": "advice_do",
    "忌": "advice_dont",
    "今日忌": "advice_dont",
    "dont": "advice_dont",
    "advice_dont": "advice_dont",
}

HEADER_ALIASES = {
    "type": "类型",
    "category": "类型",
    "text": "内容",
    "content": "内容",
    "min_rp": "RP下限",
    "rp_min": "RP下限",
    "max_rp": "RP上限",
    "rp_max": "RP上限",
    "enabled": "启用",
    "note": "备注",
}

TRUE_VALUES = {"", "1", "true", "yes", "y", "是", "启用"}
FALSE_VALUES = {"0", "false", "no", "n", "否", "禁用"}


def normalize_header(value: Any) -> str:
    text = str(value or "").replace(" ", "").strip()
    return HEADER_ALIASES.get(text.lower(), text)


def normalize_row(raw: dict[str, Any]) -> dict[str, Any]:
    return {normalize_header(key): value for key, value in raw.items() if key is not None}


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as source:
        return [normalize_row(row) for row in csv.DictReader(source)]


def read_xlsx_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("导入 .xlsx 需要 openpyxl，请先运行 pip install -r requirements.txt") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active
        values = sheet.iter_rows(values_only=True)
        try:
            headers = [normalize_header(value) for value in next(values)]
        except StopIteration:
            return []
        return [normalize_row(dict(zip(headers, row))) for row in values]
    finally:
        workbook.close()


def read_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path, sheet_name)
    raise ValueError("仅支持 .xlsx、.xlsm 或 UTF-8 CSV 文件")


def parse_enabled(value: Any, row_number: int) -> bool:
    if value is None:
        return True
    text = str(value).strip().lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    raise ValueError(f"第 {row_number} 行“启用”只能填写 是/否、1/0、true/false 或留空")


def parse_rows(rows: Iterable[dict[str, Any]]) -> dict[str, list[ContentItem]]:
    result: dict[str, list[ContentItem]] = {key: [] for key in CONTENT_KEYS}
    errors: list[str] = []
    for row_number, row in enumerate(rows, start=2):
        if not any(value not in (None, "") for value in row.values()):
            continue
        try:
            if not parse_enabled(row.get("启用"), row_number):
                continue
            raw_type = str(row.get("类型") or "").strip()
            category = TYPE_ALIASES.get(raw_type.lower())
            if not category:
                raise ValueError(f"类型“{raw_type}”无效，应为 今日签/幸运色/宜/忌")
            text = str(row.get("内容") or "").strip()
            if not text:
                raise ValueError("内容不能为空")
            min_rp = parse_optional_rp(row.get("RP下限"), "RP下限")
            max_rp = parse_optional_rp(row.get("RP上限"), "RP上限")
            if min_rp is not None and max_rp is not None and min_rp > max_rp:
                raise ValueError("RP 下限不能大于上限")
            result[category].append(
                ContentItem(
                    text=text,
                    min_rp=min_rp,
                    max_rp=max_rp,
                    note=str(row.get("备注") or "").strip(),
                )
            )
        except ValueError as exc:
            errors.append(f"第 {row_number} 行：{exc}")
    if errors:
        raise ValueError("\n".join(errors))
    return result


def validate_coverage(store: ContentStore) -> None:
    errors: list[str] = []
    for category in CONTENT_KEYS:
        missing = [rp for rp in range(101) if not store.eligible(category, rp)]
        if missing:
            ranges: list[str] = []
            start = previous = missing[0]
            for value in missing[1:]:
                if value != previous + 1:
                    ranges.append(str(start) if start == previous else f"{start}-{previous}")
                    start = value
                previous = value
            ranges.append(str(start) if start == previous else f"{start}-{previous}")
            errors.append(f"{category} 缺少可用内容的 RP 范围：{', '.join(ranges)}")
    if errors:
        raise ValueError("\n".join(errors))


def build_payload(store: ContentStore, source_name: str) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "description": "min_rp/max_rp 为 null 时表示不限制；两者均包含边界值。",
        "imported_from": source_name,
        "categories": {
            "fortune_texts": "今日签",
            "colors": "幸运色",
            "advice_do": "宜",
            "advice_dont": "忌",
        },
        **{key: [item.to_json() for item in store.data[key]] for key in CONTENT_KEYS},
    }


def write_json_atomic(payload: dict[str, Any], output_path: Path, backup: bool = True) -> Path | None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    backup_path = None
    if backup and output_path.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = output_path.with_name(f"{output_path.stem}.{timestamp}.bak{output_path.suffix}")
        shutil.copy2(output_path, backup_path)
    text = json.dumps(payload, ensure_ascii=False, indent=2) + "\n"
    with tempfile.NamedTemporaryFile(
        mode="w", encoding="utf-8", newline="\n", delete=False, dir=output_path.parent, suffix=".tmp"
    ) as temp_file:
        temp_file.write(text)
        temp_path = Path(temp_file.name)
    temp_path.replace(output_path)
    return backup_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="导入 taiko_rp 今日签、幸运色与宜忌内容表")
    parser.add_argument("input", type=Path, help=".xlsx/.xlsm 或 UTF-8 .csv 文件")
    parser.add_argument(
        "--output",
        type=Path,
        default=PROJECT_ROOT / "resource" / "content.json",
        help="输出 JSON；默认覆盖 resource/content.json 并先备份",
    )
    parser.add_argument("--sheet", default="内容库", help="Excel 工作表名称，默认“内容库”")
    parser.add_argument("--dry-run", action="store_true", help="只校验，不写入文件")
    parser.add_argument("--no-backup", action="store_true", help="覆盖时不保留带时间戳的备份")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    input_path = args.input.resolve()
    if not input_path.exists():
        print(f"错误：输入文件不存在：{input_path}", file=sys.stderr)
        return 2
    try:
        rows = read_rows(input_path, args.sheet)
        data = parse_rows(rows)
        store = ContentStore(data)
        validate_coverage(store)
        payload = build_payload(store, input_path.name)
    except (OSError, RuntimeError, ValueError) as exc:
        print(f"导入校验失败：\n{exc}", file=sys.stderr)
        return 1

    counts = {key: len(store.data[key]) for key in CONTENT_KEYS}
    print(f"校验通过：{counts}")
    if args.dry_run:
        print("dry-run：未写入任何文件。")
        return 0
    output_path = args.output.resolve()
    backup_path = write_json_atomic(payload, output_path, backup=not args.no_backup)
    print(f"已写入：{output_path}")
    if backup_path:
        print(f"原文件备份：{backup_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
